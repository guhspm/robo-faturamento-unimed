# -*- coding: utf-8 -*-

import os, re, shutil, warnings, hashlib, json, datetime, unicodedata, traceback
import pandas as pd
import numpy as np
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict, deque
from bisect import bisect_left, bisect_right

warnings.simplefilter(action='ignore', category=FutureWarning)

# ===== Configuração =====
DEBUG = True
GERAR_AUDITORIA = True
FORCAR_REEXTRACAO = True
ENTIDADES_APENAS = None  # ex.: ['HOB']

AUDITAR_PAGAMENTO = False
TOLERANCIA = Decimal("0.05")

# ⚙️ Configure os caminhos antes de executar
PASTA_MES_ANTERIOR = r"C:\caminho\para\mes_anterior"
PASTA_MES_ATUAL    = r"C:\caminho\para\mes_atual"

MESES_PT = [
    "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
    "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
]


def _meses_por_pasta(pasta_mes_atual: str):
    base = os.path.basename(pasta_mes_atual.rstrip("\\/"))
    try:
        if re.match(r'^\d{2}\.\d{4}$', base):
            m, _ = map(int, base.split("."))
            mes_atual_nome = MESES_PT[m - 1]
            m_ant = 12 if m == 1 else m - 1
            return MESES_PT[m_ant - 1], mes_atual_nome
    except Exception:
        pass
    hoje = datetime.date.today()
    m_at = hoje.month - 1 or 12
    m_an = 12 if m_at == 1 else m_at - 1
    return MESES_PT[m_an - 1], MESES_PT[m_at - 1]


MES_ANTERIOR_NOME, MES_ATUAL_NOME = _meses_por_pasta(PASTA_MES_ATUAL)
VENCIMENTO_PADRAO = "12/12/2099"


# ===== Utilitários =====
def to_dec(x) -> Decimal:
    if x is None:
        return Decimal("0")
    return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def normalize_pdf_text(txt: str):
    txt = txt.replace("\xa0", " ").replace("\u200b", "").replace("\r", "\n")
    txt = re.sub(r"[\t]+", " ", txt)
    flat = re.sub(r"\s+", " ", txt)
    return txt, flat


def strip_accents(u: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', u) if not unicodedata.combining(c))


def norm_compact(u: str) -> str:
    u = strip_accents(u).upper()
    return re.sub(r'[^A-Z0-9]', '', u)


def br_to_float(s: str) -> float:
    return float(s.replace('.', '').replace(',', '.'))


def file_md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def _cache_ruim(dados: dict) -> bool:
    return (not dados) or (dados.get('numero_nota') in (None, '', '-'))


# ===== Expressões regulares =====
HYPH = r"[\-\u2010-\u2015\u2212]"
RX_MONEY = r'(?P<valor>\d{1,3}(?:[.\s]\d{3})*,\d{2})'
RX_TITULO = re.compile(r'(?is)\bT[ií]tulo\b[:\s\-]*?(?P<titulo>[A-Z0-9]{6,20})')
RX_VALOR_TOTAL = re.compile(r'(?is)\bVALOR\s*TOTAL\s*DA\s*NFS' + HYPH + r'?\s*E\b.*?R\$\s*' + RX_MONEY)
RX_VALOR_SERV = re.compile(r'(?is)\bValor\s*do\s*Servi[cç]o\b.*?R\$\s*' + RX_MONEY)
RX_CONTRATO = re.compile(r'(?is)\bCONTRATO:\s*0*(?P<contrato>\d{6,9})\b')
RX_TIPO_COPART = re.compile(r'(?is)COPARTIC')

RX_APROX = re.compile(
    r'Valor.*?Tributos.*?ISS\s*R\$\s*(\d{1,3}(?:[.\s]\d{3})*,\d{2}).*?'
    r'PIS[/-]?\s*COFINS\s*R\$\s*(\d{1,3}(?:[.\s]\d{3})*,\d{2})',
    re.IGNORECASE | re.DOTALL
)

RX_BLOCO_FED = re.compile(
    r'TRIBUTA[ÇC][AÃ]O\s*FEDERAL(.*?)(?=VALOR\s*TOTAL\s*DA\s*NFS-?E|TOTAIS\s*APROXIMADOS|INFORMA[ÇC][ÕO]ES)',
    re.IGNORECASE | re.DOTALL
)
RX_IRRF_LINHA = re.compile(r'IRRF[\s\S]{0,200}?R\$\s*(\d{1,3}(?:[.\s]\d{3})*,\d{2})', re.IGNORECASE)
RX_TOT_RET_FED = re.compile(
    r'Total\s*das\s*Reten[cç][õo]es\s*Federais[\s\S]{0,200}?\n\s*R\$\s*(\d{1,3}(?:[.\s]\d{3})*,\d{2})',
    re.IGNORECASE
)


# ===== Índice de notas fiscais =====
class NotasIndex:
    """
    Casamento de NFs por (tipo, contrato, valor).
    Garante que cada NF seja usada no máximo uma vez e respeita o tipo da linha.
    """

    def __init__(self, notas_dict: dict[float, dict]):
        self._all = [n for _, n in notas_dict.items()]
        self._used = set()

        buckets = defaultdict(list)
        for n in self._all:
            tipo = (n.get('tipo') or 'MENSALIDADE').upper()
            contrato = n.get('contrato')
            if contrato is None:
                continue
            buckets[(tipo, int(contrato))].append(n)

        self.buckets = {}
        for key, notas in buckets.items():
            vm = defaultdict(deque)
            for n in notas:
                vm[float(n['valor_bruto'])].append(n)
            self.buckets[key] = {'vals': sorted(vm.keys()), 'map': vm}

        self.fallback_vals = []
        self.fallback_map = defaultdict(deque)
        for n in self._all:
            v = float(n['valor_bruto'])
            self.fallback_vals.append(v)
            self.fallback_map[v].append(n)
        self.fallback_vals.sort()

        by_ct_type = defaultdict(lambda: defaultdict(list))
        for n in self._all:
            ct = n.get('contrato')
            if ct is None:
                continue
            typ = (n.get('tipo') or 'MENSALIDADE').upper()
            by_ct_type[int(ct)][typ].append(n)

        self.by_contract_type = {}
        for ct, typ_map in by_ct_type.items():
            self.by_contract_type[ct] = {}
            for typ, notas in typ_map.items():
                vm = defaultdict(deque)
                for n in notas:
                    vm[float(n['valor_bruto'])].append(n)
                self.by_contract_type[ct][typ] = {'vals': sorted(vm.keys()), 'map': vm}

    def _is_used(self, n) -> bool:
        return id(n) in self._used

    def _mark_used(self, n):
        self._used.add(id(n))

    def _pop_first_not_used(self, dq: deque):
        keep = deque()
        got = None
        while dq:
            n = dq.popleft()
            if self._is_used(n):
                continue
            got = n
            break
        while dq:
            keep.append(dq.popleft())
        while keep:
            dq.appendleft(keep.pop())
        return got

    def _pop_by_title(self, key, titulo: str):
        if not titulo or key not in self.buckets:
            return None
        vals = self.buckets[key]['vals']
        m = self.buckets[key]['map']
        for v in list(vals):
            dq = m.get(v)
            if not dq:
                continue
            tmp = deque()
            got = None
            while dq:
                n = dq.popleft()
                if self._is_used(n):
                    continue
                if (n.get('titulo') or '').upper() == (titulo or '').upper():
                    got = n
                    break
                tmp.append(n)
            while dq:
                tmp.append(dq.popleft())
            m[v] = tmp
            if got is not None:
                if not m[v]:
                    try:
                        vals.remove(v)
                    except ValueError:
                        pass
                self._mark_used(got)
                return got
            if not m[v]:
                try:
                    vals.remove(v)
                except ValueError:
                    pass
        return None

    def _pop_by_value(self, idx, alvo: Decimal, tol: Decimal):
        vals = idx['vals']
        m = idx['map']
        if not vals:
            return None
        lo = float(alvo - tol)
        hi = float(alvo + tol)
        li = bisect_left(vals, lo)
        ri = bisect_right(vals, hi)
        if li >= ri:
            return None
        for v in sorted(vals[li:ri], key=lambda x: abs(Decimal(str(x)) - alvo)):
            dq = m.get(v)
            if not dq:
                continue
            n = self._pop_first_not_used(dq)
            if n is not None:
                if not dq:
                    try:
                        vals.remove(v)
                    except ValueError:
                        pass
                self._mark_used(n)
                return n
            if not dq:
                try:
                    vals.remove(v)
                except ValueError:
                    pass
        return None

    def remaining_by_contract_type(self, contrato_key: int, tipo_key: str) -> int:
        dtyp = self.by_contract_type.get(contrato_key, {}).get(tipo_key, None)
        if not dtyp:
            return 0
        total = 0
        for v in dtyp['vals']:
            dq = dtyp['map'].get(v, deque())
            total += sum(1 for n in dq if not self._is_used(n))
        return total

    def pop_any_by_contract_type(self, contrato_key: int, tipo_key: str):
        dtyp = self.by_contract_type.get(contrato_key, {}).get(tipo_key, None)
        if not dtyp:
            return None
        for v in list(dtyp['vals']):
            dq = dtyp['map'].get(v)
            if not dq:
                continue
            n = self._pop_first_not_used(dq)
            if n is not None:
                if not dq:
                    try:
                        dtyp['vals'].remove(v)
                    except ValueError:
                        pass
                self._mark_used(n)
                return n
            if not dq:
                try:
                    dtyp['vals'].remove(v)
                except ValueError:
                    pass
        return None

    def remaining_by_contract(self, contrato_key: int) -> int:
        return (
            self.remaining_by_contract_type(contrato_key, 'MENSALIDADE')
            + self.remaining_by_contract_type(contrato_key, 'COPART')
        )

    def pop_any_by_contract(self, contrato_key: int):
        n = self.pop_any_by_contract_type(contrato_key, 'MENSALIDADE')
        if n is not None:
            return n
        return self.pop_any_by_contract_type(contrato_key, 'COPART')

    def pop_match(self, tipo_linha: str, contrato_linha, alvo: Decimal, titulo_linha: str = None):
        tipo_key = (
            'COPART'
            if ('COPART' in (tipo_linha or '').upper() or 'COPARTIC' in (tipo_linha or '').upper())
            else 'MENSALIDADE'
        )
        try:
            contrato_key = int(str(contrato_linha).lstrip('0') or 0)
        except Exception:
            contrato_key = contrato_linha
        key = (tipo_key, contrato_key)

        if key in self.buckets:
            count = sum(
                len([n for n in dq if not self._is_used(n)])
                for dq in self.buckets[key]['map'].values()
            )
            if count == 1:
                for v in list(self.buckets[key]['vals']):
                    dq = self.buckets[key]['map'].get(v)
                    if not dq:
                        continue
                    n = self._pop_first_not_used(dq)
                    if n:
                        if not dq:
                            try:
                                self.buckets[key]['vals'].remove(v)
                            except ValueError:
                                pass
                        self._mark_used(n)
                        return n

        n = self._pop_by_title(key, titulo_linha)
        if n:
            return n

        if key in self.buckets:
            n = self._pop_by_value(self.buckets[key], alvo, TOLERANCIA)
            if n:
                return n

        tol2 = max((alvo * Decimal("0.005")), Decimal("3.00"))
        if key in self.buckets:
            n = self._pop_by_value(self.buckets[key], alvo, tol2)
            if n:
                return n

        n = self._pop_global_tipo(alvo, TOLERANCIA, tipo_key)
        if n:
            return n
        n = self._pop_global_tipo(alvo, tol2, tipo_key)
        if n:
            return n

        return None

    def _pop_global_tipo(self, alvo: Decimal, tol: Decimal, tipo_key: str):
        vals = self.fallback_vals
        m = self.fallback_map
        if not vals:
            return None
        lo = float(alvo - tol)
        hi = float(alvo + tol)
        li = bisect_left(vals, lo)
        ri = bisect_right(vals, hi)
        if li >= ri:
            return None
        for v in sorted(vals[li:ri], key=lambda x: abs(Decimal(str(x)) - alvo)):
            dq = m.get(v)
            if not dq:
                continue
            tmp = deque()
            got = None
            while dq:
                n = dq.popleft()
                if self._is_used(n):
                    continue
                if (n.get('tipo') or 'MENSALIDADE').upper() == tipo_key:
                    got = n
                    break
                tmp.append(n)
            while dq:
                tmp.append(dq.popleft())
            m[v] = tmp
            if got is not None:
                if not m[v]:
                    try:
                        vals.remove(v)
                    except ValueError:
                        pass
                self._mark_used(got)
                return got
            if not m[v]:
                try:
                    vals.remove(v)
                except ValueError:
                    pass
        return None


# ===== Extração de PDFs =====
def extrair_dados_pdfs(caminho_entidade: str, usar_cache=None) -> dict[float, dict]:
    if usar_cache is None:
        usar_cache = not FORCAR_REEXTRACAO

    dicionario_nfs: dict[float, list] = defaultdict(list)
    pdfs = []
    for root, _, files in os.walk(caminho_entidade):
        for f in files:
            if f.lower().endswith('.pdf'):
                pdfs.append(os.path.join(root, f))
    if not pdfs:
        return {}

    for caminho_pdf in pdfs:
        try:
            cache_path = caminho_pdf + '.json'
            cur_hash = file_md5(caminho_pdf)
            usar_cache_arquivo = usar_cache
            dados_cache = None

            if usar_cache_arquivo and os.path.exists(cache_path):
                try:
                    cache = json.load(open(cache_path, 'r', encoding='utf-8'))
                    if cache.get('md5') == cur_hash:
                        dados_cache = cache.get('dados', {})
                    if _cache_ruim(dados_cache):
                        usar_cache_arquivo = False
                except Exception:
                    usar_cache_arquivo = False

            if usar_cache_arquivo and dados_cache:
                dados = dados_cache
            else:
                with pdfplumber.open(caminho_pdf) as pdf:
                    texto_cru = ""
                    for p in pdf.pages:
                        t = p.extract_text() or ""
                        texto_cru += t + "\n"

                texto, flat = normalize_pdf_text(texto_cru)
                dados = {
                    'numero_nota': '-', 'titulo': '',
                    'valor_bruto': 0.0, 'valor_servico': 0.0,
                    'iss': 0.0, 'irrf': 0.0, 'pis_cofins': 0.0,
                    'contrato': None, 'tipo': None, 'vencimento': None,
                    'arquivo_pdf': caminho_pdf
                }

                linhas = texto.splitlines()
                for i, ln in enumerate(linhas):
                    lc = norm_compact(ln)
                    if "NUMERODANFSE" in lc:
                        m_here = re.search(r'\b(\d{4,9})\b', ln)
                        if m_here:
                            dados['numero_nota'] = m_here.group(1)
                            break
                        for j in (1, 2):
                            if i + j < len(linhas):
                                m_next = re.search(r'\b(\d{4,9})\b', linhas[i + j])
                                if m_next:
                                    dados['numero_nota'] = m_next.group(1)
                                    break
                    if dados['numero_nota'] != '-':
                        break

                m = RX_TITULO.search(texto) or RX_TITULO.search(flat)
                if m:
                    dados['titulo'] = m.group('titulo')
                m = RX_CONTRATO.search(texto) or RX_CONTRATO.search(flat)
                if m:
                    try:
                        dados['contrato'] = int(m.group('contrato'))
                    except Exception:
                        pass
                dados['tipo'] = (
                    'COPART'
                    if (RX_TIPO_COPART.search(texto) or RX_TIPO_COPART.search(flat))
                    else 'MENSALIDADE'
                )

                bruto = None
                m = RX_VALOR_TOTAL.search(texto) or RX_VALOR_TOTAL.search(flat)
                if not m:
                    m = RX_VALOR_SERV.search(texto) or RX_VALOR_SERV.search(flat)
                if m:
                    bruto = br_to_float(m.group('valor'))
                else:
                    valores = re.findall(r'\d{1,3}(?:[.\s]\d{3})*,\d{2}', flat)
                    if valores:
                        bruto = max(br_to_float(v) for v in valores)
                if bruto is None:
                    raise ValueError("Valor bruto não identificado no PDF.")
                dados['valor_bruto'] = round(bruto, 2)
                dados['valor_servico'] = round(bruto, 2)

                m_ap = RX_APROX.search(flat)
                if m_ap:
                    dados['iss'] = round(br_to_float(m_ap.group(1)), 2)
                    dados['pis_cofins'] = round(br_to_float(m_ap.group(2)), 2)

                m_blk = RX_BLOCO_FED.search(flat)
                m_irrf = RX_IRRF_LINHA.search(m_blk.group(1)) if m_blk else None
                if not m_irrf:
                    m_irrf = RX_IRRF_LINHA.search(flat)
                if m_irrf and m_irrf.group(1):
                    dados['irrf'] = round(br_to_float(m_irrf.group(1)), 2)
                else:
                    m_tot = RX_TOT_RET_FED.search(texto) or RX_TOT_RET_FED.search(flat)
                    dados['irrf'] = round(br_to_float(m_tot.group(1)), 2) if (m_tot and m_tot.group(1)) else 0.0

                RX_VENC = re.compile(r'(?is)\bVencimento\s*:\s*(?P<ven>\d{2}/\d{2}/\d{4})')
                m = RX_VENC.search(texto) or RX_VENC.search(flat)
                if m:
                    dados['vencimento'] = m.group('ven')

                if DEBUG:
                    print(
                        f"[PDF] {os.path.basename(caminho_pdf)} -> "
                        f"NF {dados.get('numero_nota', '-')} "
                        f"Título {dados.get('titulo', '')} "
                        f"Bruto R$ {dados['valor_bruto']:.2f} "
                        f"ISS R$ {dados['iss']:.2f} "
                        f"IR R$ {dados['irrf']:.2f} "
                        f"PIS/COFINS R$ {dados['pis_cofins']:.2f} "
                        f"Tipo {dados.get('tipo', '?')} "
                        f"Contrato {dados.get('contrato')} "
                        f"Venc {dados.get('vencimento') or '-'}"
                    )

                dicionario_nfs[dados['valor_bruto']].append(dados)

                try:
                    with open(cache_path, 'w', encoding='utf-8') as w:
                        json.dump({'md5': cur_hash, 'dados': dados}, w, ensure_ascii=False, indent=2)
                except Exception:
                    pass

        except Exception as e:
            print(f"[ERRO PDF] {os.path.basename(caminho_pdf)}: {e}")

    return {k: v[0] for k, v in dicionario_nfs.items()}


# ===== Processamento Excel =====
CREDIT_TYPES = [6, 9, 10, 11, 15, 16, 17, 21]


def processar_bases_excel(caminho_excel: str) -> pd.DataFrame:
    dfs = pd.read_excel(caminho_excel, sheet_name=None, engine='openpyxl')
    df_mensa = dfs.get('MENSALIDADES')
    df_copart = dfs.get('COPART')
    if df_mensa is None or df_copart is None:
        raise RuntimeError("Abas 'MENSALIDADES' e/ou 'COPART' não encontradas no bruto.")

    df_mensa['Origem'] = 'MENSALIDADE'
    df_copart['Origem'] = 'COPART'
    df_base = pd.concat([df_mensa, df_copart], ignore_index=True)

    tipos_ativos = [0, 1, 3, 5, 12, 13, 14, 18, 19, 20]
    df_ativos = df_base[df_base['Tipo_Debito'].isin(tipos_ativos)].copy()
    df_ativos['MNC'] = np.where(df_ativos['Vl_Minimo_MNC'] == 'N', df_ativos['Retorno_Vl_NC'], 0)
    df_ativos['MNC_20'] = np.where(df_ativos['Vl_Minimo_MNC'] == 'S', df_ativos['Retorno_Vl_NC'], 0)

    resumo = (
        df_ativos
        .groupby(['Codigo_Contrato', 'Origem'], as_index=False)
        .agg(
            sub=('Retorno_Vl_Subsidio', 'sum'),
            mcf=('Retorno_Vl_MC', 'sum'),
            mnc=('MNC', 'sum'),
            mnc20=('MNC_20', 'sum'),
        )
    ).round(2)

    df_cred = df_base[df_base['Tipo_Debito'].isin(CREDIT_TYPES)].copy()
    for c in ('Retorno_Vl_Subsidio', 'Retorno_Vl_MC'):
        if c in df_cred.columns:
            df_cred[c] = pd.to_numeric(df_cred[c], errors='coerce').fillna(0.0)
    cred = (
        df_cred
        .groupby(['Codigo_Contrato', 'Origem'], as_index=False)
        .agg(cred_sub=('Retorno_Vl_Subsidio', 'sum'), cred_mcf=('Retorno_Vl_MC', 'sum'))
    ).round(2)

    resumo = resumo.merge(cred, on=['Codigo_Contrato', 'Origem'], how='left')
    for c in ('cred_sub', 'cred_mcf'):
        if c not in resumo.columns:
            resumo[c] = 0.0
    resumo[['cred_sub', 'cred_mcf']] = resumo[['cred_sub', 'cred_mcf']].fillna(0.0).round(2)
    return resumo


def indexar_calculo(df_calc: pd.DataFrame):
    out = defaultdict(dict)
    for _, r in df_calc.iterrows():
        contrato = int(r['Codigo_Contrato'])
        origem = r['Origem']
        out[contrato][origem] = {
            'sub': float(r.get('sub', 0) or 0),
            'mcf': float(r.get('mcf', 0) or 0),
            'mnc': float(r.get('mnc', 0) or 0),
            'mnc20': float(r.get('mnc20', 0) or 0),
            'cred_sub': float(r.get('cred_sub', 0) or 0),
            'cred_mcf': float(r.get('cred_mcf', 0) or 0),
        }
    return out


# ===== Localização de cabeçalho =====
def localizar_cabecalho(ws, chaves=('TIPO', 'CONTRATO')):
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [str(c.value).strip().upper() if c.value is not None else '' for c in ws[r]]
        if 'TIPO' in row_vals and 'CONTRATO' in row_vals:
            headers = {row_vals[i]: i + 1 for i in range(len(row_vals)) if row_vals[i]}
            return r, headers
    raise RuntimeError(f'Cabeçalho não encontrado em "{ws.title}"')


def _col(H: dict, name: str, aliases=()):
    nameU = name.upper()
    if nameU in H:
        return H[nameU]
    for a in aliases:
        if a.upper() in H:
            return H[a.upper()]
    if 'MNC' in nameU and '20' in nameU:
        for k in H:
            if 'MNC' in k and '20' in k:
                return H[k]
    return None


# ===== Preenchimento do layout =====
def preencher_layout(wb_path: str, df_calc: pd.DataFrame, notas_dict: dict,
                     mes_anterior: str, mes_atual: str, vencimento_padrao: str):
    wb = openpyxl.load_workbook(wb_path, data_only=False, keep_links=True)
    calc_idx = indexar_calculo(df_calc)
    notas_idx_nf = NotasIndex(notas_dict)
    notas_idx_pay = NotasIndex(notas_dict)

    for aba in ('DADOS DA NF', 'DADOS PARA PAGAMENTO'):
        if aba not in wb.sheetnames:
            print(f"[AVISO] Aba '{aba}' não encontrada no modelo.")
            continue
        ws = wb[aba]

        for r in range(1, min(6, ws.max_row) + 1):
            for c in range(1, min(6, ws.max_column) + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and mes_anterior in v:
                    ws.cell(r, c).value = v.replace(mes_anterior, mes_atual)

        hdr_row, H = localizar_cabecalho(ws, ('TIPO', 'CONTRATO'))
        start = hdr_row + 1

        r = start
        while r <= ws.max_row:
            jContrato = _col(H, 'CONTRATO') or 2
            contrato = ws.cell(r, jContrato).value
            if not contrato or str(contrato).strip() in ('', '-'):
                break

            if aba == 'DADOS DA NF':
                for nome, aliases in (
                    ('Nº NOTA', ('N° NOTA', 'Nº DA NOTA')),
                    ('TITULO', ('TÍTULO', 'TITULO NF')),
                    ('SUBSIDIO', ()),
                    ('MCF', ()),
                    ('CREDITOS', ('CRÉDITOS',)),
                    ('MNC', ()),
                    ('MNC < 20,00', ('MNC < 20', 'MNC<20', 'MNC \\u003c 20,00')),
                    ('VALOR BRUTO', ()),
                ):
                    j = _col(H, nome, aliases)
                    if not j:
                        continue
                    if nome in ('Nº NOTA', 'TITULO'):
                        ws.cell(r, j).value = '-'
                    else:
                        ws.cell(r, j).value = 0
                jv = _col(H, 'VENCIMENTO')
                if jv:
                    ws.cell(r, jv).value = '-'
            else:
                SKIP = {'RETENÇÃO DE 0,25%', 'RETENCAO DE 0,25%'}
                for nome, aliases in (
                    ('Nº NOTA', ('N° NOTA', 'Nº DA NOTA')),
                    ('SUBSIDIO', ()),
                    ('MARGEM CONSIGNADA', ('MCF', 'MARGEM CONSIG.')),
                    ('CREDITOS', ('CRÉDITOS',)),
                    ('RETENÇÃO DE 0,25%', ('RETENCAO DE 0,25%',)),
                    ('ISS', ('ISSQN',)),
                    ('IR', ('IRRF', 'IR RF', 'IRRF ')),
                    ('PIS/COFINS', ('PIS-COFINS', 'PIS COFINS')),
                ):
                    if nome in SKIP:
                        continue
                    j = _col(H, nome, aliases)
                    if not j:
                        continue
                    if nome == 'Nº NOTA':
                        ws.cell(r, j).value = '-'
                    else:
                        ws.cell(r, j).value = 0
            r += 1

        r = start
        while r <= ws.max_row:
            jTipo = _col(H, 'TIPO') or 1
            jContrato = _col(H, 'CONTRATO') or 2
            contrato = ws.cell(r, jContrato).value
            if not contrato or str(contrato).strip() in ('', '-'):
                break

            tipo_raw = ws.cell(r, jTipo).value
            tipo_up = (str(tipo_raw) if tipo_raw else '').upper()
            tipo_norm = strip_accents(tipo_up)
            tipo_key = 'COPART' if ('COPART' in tipo_norm or 'COPARTIC' in tipo_norm) else 'MENSALIDADE'

            is_copart = ('COPART' in tipo_norm) or ('COPARTIC' in tipo_norm)
            is_subsidio = ('SUBSIDIO' in tipo_norm)

            v = {'sub': 0.0, 'mcf': 0.0, 'mnc': 0.0, 'mnc20': 0.0, 'cred_sub': 0.0, 'cred_mcf': 0.0}
            try:
                contrato_key = int(str(contrato).lstrip('0') or 0)
            except Exception:
                contrato_key = contrato
            calculo = calc_idx.get(contrato_key, {})

            if is_copart and is_subsidio:
                base = calculo.get('COPART', {})
                v['sub'] = base.get('sub', 0.0)
                v['cred_sub'] = base.get('cred_sub', 0.0)
                v['mcf'] = v['mnc'] = v['mnc20'] = 0.0
                v['cred_mcf'] = 0.0
            elif is_copart:
                base = calculo.get('COPART', {})
                v['mcf'] = base.get('mcf', 0.0)
                v['mnc'] = base.get('mnc', 0.0)
                v['mnc20'] = base.get('mnc20', 0.0)
                v['cred_sub'] = base.get('cred_sub', 0.0)
                v['cred_mcf'] = base.get('cred_mcf', 0.0)
            elif 'SUBSIDIO' in tipo_norm:
                base = calculo.get('MENSALIDADE', {})
                v['sub'] = base.get('sub', 0.0)
                v['cred_sub'] = base.get('cred_sub', 0.0)
            elif 'MCF' in tipo_up:
                base = calculo.get('MENSALIDADE', {})
                v['mcf'] = base.get('mcf', 0.0)
                v['mnc'] = base.get('mnc', 0.0)
                v['mnc20'] = base.get('mnc20', 0.0)
                v['cred_mcf'] = base.get('cred_mcf', 0.0)

            cred_comp = to_dec(v['cred_sub']) if is_subsidio else to_dec(v['cred_mcf'])
            tot = to_dec(v['sub']) + to_dec(v['mcf']) + to_dec(v['mnc']) + to_dec(v['mnc20']) - cred_comp

            j_tit = _col(H, 'TITULO', ('TÍTULO', 'TITULO NF'))
            titulo_da_linha = (
                ws.cell(r, j_tit).value
                if (j_tit and ws.cell(r, j_tit).value and ws.cell(r, j_tit).value != '-')
                else None
            )

            if aba == 'DADOS DA NF':
                j_sub = _col(H, 'SUBSIDIO')
                j_mcf = _col(H, 'MCF')
                j_mnc = _col(H, 'MNC')
                j_mnc2 = _col(H, 'MNC < 20,00', ('MNC < 20', 'MNC<20', 'MNC \\u003c 20,00'))
                j_cred = _col(H, 'CREDITOS', ('CRÉDITOS',))
                j_nota = _col(H, 'Nº NOTA', ('N° NOTA', 'Nº DA NOTA'))
                j_venc = _col(H, 'VENCIMENTO')
                j_vb = _col(H, 'VALOR BRUTO')

                if j_sub: ws.cell(r, j_sub).value = float(v['sub'])
                if j_mcf: ws.cell(r, j_mcf).value = float(v['mcf'])
                if j_mnc: ws.cell(r, j_mnc).value = float(v['mnc'])
                if j_mnc2: ws.cell(r, j_mnc2).value = float(v['mnc20'])

                if j_cred:
                    ws.cell(r, j_cred).value = (
                        -float(v.get('cred_sub', 0.0)) if is_subsidio else -float(v.get('cred_mcf', 0.0))
                    )

                if j_vb:
                    from openpyxl.utils import get_column_letter
                    cols = [c for c in (j_sub, j_mcf, j_cred, j_mnc, j_mnc2) if c]
                    if cols:
                        c1 = get_column_letter(min(cols))
                        c2 = get_column_letter(max(cols))
                        ws.cell(r, j_vb).value = f"=SUM({c1}{r}:{c2}{r})"
                        ws.cell(r, j_vb).number_format = 'R$ #,##0.00'

                nota = None
                try:
                    nota = notas_idx_nf.pop_match(tipo_up, contrato, tot, titulo_da_linha)
                except Exception as e:
                    if DEBUG:
                        print(f"[AVISO] Match NF falhou: {e}")

                if (nota is None) and (float(tot) == 0.0):
                    try:
                        ct_key = int(str(contrato).lstrip('0') or 0)
                    except Exception:
                        ct_key = contrato
                    if notas_idx_nf.remaining_by_contract_type(ct_key, tipo_key) == 1:
                        nota = notas_idx_nf.pop_any_by_contract_type(ct_key, tipo_key)
                    if nota:
                        val_pdf = float(nota.get('valor_servico') or nota.get('valor_bruto') or 0.0)
                        if ('MCF' in tipo_up) and j_mcf:
                            ws.cell(r, j_mcf).value = val_pdf
                        elif ('SUBSIDIO' in tipo_norm) and j_sub:
                            ws.cell(r, j_sub).value = val_pdf
                        elif j_mcf:
                            ws.cell(r, j_mcf).value = val_pdf
                        elif j_sub:
                            ws.cell(r, j_sub).value = val_pdf
                        if j_cred:
                            ws.cell(r, j_cred).value = 0.0

                if nota:
                    if j_nota: ws.cell(r, j_nota).value = nota.get('numero_nota', '-')
                    if j_tit: ws.cell(r, j_tit).value = nota.get('titulo', '-')
                    if j_venc and nota.get('vencimento'):
                        ws.cell(r, j_venc).value = nota['vencimento']

            else:
                j_sub = _col(H, 'SUBSIDIO')
                j_mcf = _col(H, 'MARGEM CONSIGNADA', ('MCF', 'MARGEM CONSIG.'))
                j_cred = _col(H, 'CREDITOS', ('CRÉDITOS',))
                j_nota = _col(H, 'Nº NOTA', ('N° NOTA', 'Nº DA NOTA'))
                j_iss = _col(H, 'ISS', ('ISSQN',))
                j_ir = _col(H, 'IR', ('IRRF', 'IR RF', 'IRRF '))
                j_pis = _col(H, 'PIS/COFINS', ('PIS-COFINS', 'PIS COFINS'))

                if j_sub: ws.cell(r, j_sub).value = float(v['sub'])
                if j_mcf: ws.cell(r, j_mcf).value = float(v['mcf'])
                if j_cred:
                    ws.cell(r, j_cred).value = (
                        -float(v.get('cred_sub', 0.0)) if is_subsidio else -float(v.get('cred_mcf', 0.0))
                    )

                nota = None
                try:
                    nota = notas_idx_pay.pop_match(tipo_up, contrato, tot, titulo_da_linha)
                except Exception as e:
                    if DEBUG:
                        print(f"[AVISO] Match PAY falhou: {e}")

                if (nota is None) and (float(tot) == 0.0):
                    try:
                        ct_key = int(str(contrato).lstrip('0') or 0)
                    except Exception:
                        ct_key = contrato
                    if notas_idx_pay.remaining_by_contract_type(ct_key, tipo_key) == 1:
                        nota = notas_idx_pay.pop_any_by_contract_type(ct_key, tipo_key)

                if nota:
                    if j_nota: ws.cell(r, j_nota).value = nota.get('numero_nota', '-')
                    if j_iss: ws.cell(r, j_iss).value = -float(nota.get('iss', 0.0))
                    if j_ir: ws.cell(r, j_ir).value = -float(nota.get('irrf', 0.0))
                    if j_pis: ws.cell(r, j_pis).value = -float(nota.get('pis_cofins', 0.0))

            r += 1

    wb.save(wb_path)


# ===== Auditoria =====
def _fmt_num_br(v: float) -> str:
    return f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _banner(txt: str):
    line = "═" * max(8, len(txt) + 2)
    print("\n" + "╔" + line + "╗")
    print("║ " + txt + " ║")
    print("╚" + line + "╝")


def auditar_planilha(caminho_saida: str, nfs: dict, entidade: str):
    """
    Compara valores brutos entre PDFs e layout preenchido.

    Por NF: cruza pelo número da nota; se não casar, tenta por contrato + valor.
    Por contrato: compara somatórios de cada lado.
    """
    try:
        wb = openpyxl.load_workbook(caminho_saida, data_only=True)
        if 'DADOS DA NF' not in wb.sheetnames:
            print("[AUDITORIA] Aba DADOS DA NF não encontrada.")
            return None
        ws_nf = wb['DADOS DA NF']
    except Exception as e:
        print(f"[AUDITORIA] Falha abrindo o arquivo: {e}")
        return None

    hdr_row, H = None, {}
    for r in range(1, min(ws_nf.max_row, 40) + 1):
        row_vals = [str(c.value).strip().upper() if c.value is not None else '' for c in ws_nf[r]]
        if 'TIPO' in row_vals and 'CONTRATO' in row_vals:
            H = {row_vals[i]: i for i in range(len(row_vals)) if row_vals[i]}
            hdr_row = r
            break
    if hdr_row is None:
        print("[AUDITORIA] Cabeçalho não localizado na aba DADOS DA NF.")
        return None

    headers_nf = [str(c.value).strip() if c.value is not None else '' for c in ws_nf[hdr_row]]
    data_nf_sheet = []
    for r in ws_nf.iter_rows(min_row=hdr_row + 1, values_only=True):
        idx = H.get('CONTRATO', None)
        if idx is None:
            break
        cv = r[idx] if idx < len(r) else None
        if (cv is None) or str(cv).strip() in ('', '-'):
            break
        data_nf_sheet.append(list(r[:len(headers_nf)]))
    df_nf_sheet = pd.DataFrame(data_nf_sheet, columns=headers_nf)

    rename_nf_map = {
        'Nº NOTA ': 'Nº NOTA',
        'SUBSÍDIO': 'SUBSIDIO',
        'CRÉDITOS': 'CREDITOS',
        'CREDITOS ': 'CREDITOS',
        'MNC < 20': 'MNC < 20,00',
        'MNC<20': 'MNC < 20,00',
        'MNC \\u003c 20,00': 'MNC < 20,00'
    }
    for k, v in rename_nf_map.items():
        if k in df_nf_sheet.columns:
            df_nf_sheet.rename(columns={k: v}, inplace=True)

    for col in ['SUBSIDIO', 'MCF', 'CREDITOS', 'MNC', 'MNC < 20,00', 'VALOR BRUTO']:
        if col not in df_nf_sheet.columns:
            df_nf_sheet[col] = 0.0
        df_nf_sheet[col] = pd.to_numeric(df_nf_sheet[col], errors='coerce').fillna(0.0)

    if 'CONTRATO' in df_nf_sheet.columns:
        df_nf_sheet['CONTRATO'] = df_nf_sheet['CONTRATO'].astype(str).str.strip()
    if 'TIPO' not in df_nf_sheet.columns:
        df_nf_sheet['TIPO'] = ''

    df_nf_sheet['CONTRATO_KEY'] = (
        df_nf_sheet['CONTRATO'].astype(str)
        .str.replace(r'\D', '', regex=True)
        .str.lstrip('0')
        .replace('', '0')
    )
    df_nf_sheet['_Nº NOTA norm'] = df_nf_sheet.get('Nº NOTA', '-').astype(str)
    df_nf_sheet['NF_NORM'] = df_nf_sheet['_Nº NOTA norm'].str.replace(r'\D', '', regex=True)

    df_nf_sheet['BRUTO(layout)'] = (
        df_nf_sheet['SUBSIDIO'].fillna(0)
        + df_nf_sheet['MCF'].fillna(0)
        + df_nf_sheet['CREDITOS'].fillna(0)
        + df_nf_sheet['MNC'].fillna(0)
        + df_nf_sheet['MNC < 20,00'].fillna(0)
    )

    rows = []
    for d in nfs.values():
        contrato_raw = str(d.get('contrato') or '')
        contrato_key = re.sub(r'\D', '', contrato_raw).lstrip('0') or '0'
        nf_raw = str(d.get('numero_nota') or '-')
        nf_key = re.sub(r'\D', '', nf_raw)
        rows.append({
            'ENTIDADE': entidade,
            'CONTRATO_KEY': contrato_key,
            'Nº NOTA (pdf)': nf_key,
            'VALOR_SERVICO(pdf)': d.get('valor_servico') or d.get('valor_bruto') or 0.0,
        })
    df_pdf = pd.DataFrame(rows)
    if df_pdf.empty:
        print("[AUDITORIA] Nenhuma NF extraída para auditar.")
        return None
    df_pdf['VALOR_SERVICO(pdf)'] = pd.to_numeric(df_pdf['VALOR_SERVICO(pdf)'], errors='coerce').fillna(0.0)

    m1 = df_pdf.merge(
        df_nf_sheet[['CONTRATO_KEY', 'NF_NORM', 'BRUTO(layout)', 'TIPO']].rename(
            columns={'NF_NORM': 'Nº NOTA (pdf)'}
        ),
        how='left',
        on=['CONTRATO_KEY', 'Nº NOTA (pdf)'],
    )
    m1['CONTRATO'] = m1['CONTRATO_KEY']

    not_matched = m1['BRUTO(layout)'].isna()
    if not_matched.any():
        df_un = m1[not_matched].copy()
        df_ok = m1[~not_matched].copy()

        base_ct = df_nf_sheet[['CONTRATO_KEY', 'BRUTO(layout)', 'TIPO']].copy()
        base_ct.rename(columns={'CONTRATO_KEY': 'CONTRATO'}, inplace=True)
        base_ct['_idx_layout'] = base_ct.index

        casados = []
        for ct, gpdf in df_un.groupby('CONTRATO'):
            base = base_ct[base_ct['CONTRATO'] == ct].copy()
            usados = set()
            for i, rr in gpdf.iterrows():
                alvo = Decimal(str(rr['VALOR_SERVICO(pdf)']))
                cand = []
                for j, bb in base.iterrows():
                    if bb['_idx_layout'] in usados:
                        continue
                    dif = abs(float(Decimal(str(bb['BRUTO(layout)'])) - alvo))
                    cand.append((dif, j))
                if not cand:
                    continue
                cand.sort(key=lambda x: x[0])
                dif, jbest = cand[0]
                tol2 = max((alvo * Decimal("0.005")), Decimal("3.00"))
                if dif <= float(tol2):
                    usados.add(base.loc[jbest, '_idx_layout'])
                    casados.append((i, float(base.loc[jbest, 'BRUTO(layout)']), base.loc[jbest, 'TIPO']))

        if casados:
            idx_pdf = [i for (i, _, _) in casados]
            vb_lay = [v for (_, v, _) in casados]
            tipos = [t for (_, _, t) in casados]
            df_un.loc[idx_pdf, 'BRUTO(layout)'] = vb_lay
            df_un.loc[idx_pdf, 'TIPO'] = tipos

        m1 = pd.concat([df_ok, df_un], ignore_index=False).sort_index()

    m1['BRUTO(layout)'] = pd.to_numeric(m1['BRUTO(layout)'], errors='coerce').fillna(0.0)
    m1['DIF_BRUTO(pdf_vs_layout)'] = (
        pd.to_numeric(m1['VALOR_SERVICO(pdf)'], errors='coerce').fillna(0.0) - m1['BRUTO(layout)']
    ).round(2)

    _banner(f"AUDITORIA • {entidade} • DADOS DA NF (BRUTO por NF)")
    ok_nf = (m1['DIF_BRUTO(pdf_vs_layout)'].abs() <= float(TOLERANCIA)).sum()
    tot_nf = len(m1)
    print(f"Resumo: {ok_nf}/{tot_nf} NFs batendo. Tolerância: R$ {float(TOLERANCIA):.2f}.")
    diverg_nf = m1[m1['DIF_BRUTO(pdf_vs_layout)'].abs() > float(TOLERANCIA)]
    if diverg_nf.empty:
        print("Tudo certo ✅")
    else:
        print("\nDivergências:")
        for _, r in diverg_nf.iterrows():
            print(
                f" • Contrato {r['CONTRATO']} "
                f"NF {r['Nº NOTA (pdf)']} "
                f"PDF R$ {_fmt_num_br(r['VALOR_SERVICO(pdf)'])} "
                f"vs Layout R$ {_fmt_num_br(r['BRUTO(layout)'])} "
                f"→ Dif R$ {_fmt_num_br(r['DIF_BRUTO(pdf_vs_layout)'])} "
                f"{'✗' if abs(float(r['DIF_BRUTO(pdf_vs_layout)'])) > float(TOLERANCIA) else '✓'}"
            )

    _banner(f"AUDITORIA • {entidade} • DADOS DA NF (BRUTO por CONTRATO)")
    agg_pdf = (
        df_pdf.groupby('CONTRATO_KEY', as_index=False)['VALOR_SERVICO(pdf)']
        .sum()
        .rename(columns={'CONTRATO_KEY': 'CONTRATO', 'VALOR_SERVICO(pdf)': 'BRUTO(pdf)_contrato'})
    )
    agg_lay = (
        df_nf_sheet.groupby('CONTRATO_KEY', as_index=False)['BRUTO(layout)']
        .sum()
        .rename(columns={'CONTRATO_KEY': 'CONTRATO', 'BRUTO(layout)': 'BRUTO(layout)_contrato'})
    )
    cmpc = agg_pdf.merge(agg_lay, on='CONTRATO', how='outer').fillna(0.0)
    cmpc['DIF_CONTRATO(pdf_vs_layout)'] = (
        cmpc['BRUTO(pdf)_contrato'] - cmpc['BRUTO(layout)_contrato']
    ).round(2)
    ok_ct = (cmpc['DIF_CONTRATO(pdf_vs_layout)'].abs() <= float(TOLERANCIA)).sum()
    tot_ct = len(cmpc)
    print(f"Resumo: {ok_ct}/{tot_ct} contratos batendo. Tolerância: R$ {float(TOLERANCIA):.2f}.")
    diverg_ct = cmpc[cmpc['DIF_CONTRATO(pdf_vs_layout)'].abs() > float(TOLERANCIA)]
    if diverg_ct.empty:
        print("Tudo certo ✅")
    else:
        print("\nDivergências por contrato:")
        for _, r in diverg_ct.iterrows():
            print(
                f" • Contrato {r['CONTRATO']} "
                f"PDF Σ R$ {_fmt_num_br(r['BRUTO(pdf)_contrato'])} "
                f"vs Layout Σ R$ {_fmt_num_br(r['BRUTO(layout)_contrato'])} "
                f"→ Dif R$ {_fmt_num_br(r['DIF_CONTRATO(pdf_vs_layout)'])} "
                f"{'✗' if abs(float(r['DIF_CONTRATO(pdf_vs_layout)'])) > float(TOLERANCIA) else '✓'}"
            )

    m1_out = m1.copy()
    m1_out['ENTIDADE'] = entidade
    m1_out = m1_out[['ENTIDADE'] + [c for c in m1_out.columns if c != 'ENTIDADE']]

    cmpc_out = cmpc.copy()
    cmpc_out['ENTIDADE'] = entidade
    cmpc_out = cmpc_out[['ENTIDADE'] + [c for c in cmpc_out.columns if c != 'ENTIDADE']]

    return {
        'por_nf': m1_out,
        'por_contrato': cmpc_out,
        'ok_nf': int(ok_nf), 'tot_nf': int(tot_nf),
        'ok_ct': int(ok_ct), 'tot_ct': int(tot_ct),
    }


# ===== Orquestração =====
def executar():
    p_ant = PASTA_MES_ANTERIOR
    p_at = PASTA_MES_ATUAL
    if not os.path.isdir(p_at):
        print(f"[ERRO] Pasta do mês atual não encontrada: {p_at}")
        return []
    if not os.path.isdir(p_ant):
        print(f"[ERRO] Pasta do mês anterior não encontrada: {p_ant}")
        return []

    entidades = [d for d in os.listdir(p_at) if os.path.isdir(os.path.join(p_at, d))]
    if ENTIDADES_APENAS:
        entidades = [e for e in entidades if e in ENTIDADES_APENAS]
    print(f"[INFO] Entidades detectadas em {p_at}: {entidades}")

    relatorios = []
    for ent in entidades:
        try:
            print(f"\n-> Entidade: {ent}")
            path_ent_at = os.path.join(p_at, ent)
            path_ent_ant = os.path.join(p_ant, ent)

            f_bruto = [
                f for f in os.listdir(path_ent_at)
                if f.lower().endswith('.xlsx') and '_pronto' not in f.lower() and not f.startswith('~$')
            ]
            print(f"  [INFO] Brutos encontrados (atual): {f_bruto}")
            if not f_bruto:
                print("  [PULA] Sem bruto .xlsx válido na pasta.")
                continue

            if not os.path.isdir(path_ent_ant):
                print("  [PULA] Pasta da entidade no mês anterior não existe.")
                continue

            f_molde = [
                f for f in os.listdir(path_ent_ant)
                if f.lower().endswith('.xlsx') and not f.startswith('~$')
            ]
            print(f"  [INFO] Moldes encontrados (mês anterior): {f_molde}")
            if not f_molde:
                print("  [PULA] Sem molde .xlsx no mês anterior.")
                continue

            nfs = extrair_dados_pdfs(path_ent_at)
            caminho_bruto = os.path.join(path_ent_at, f_bruto[0])
            print(f"  [INFO] Usando bruto: {caminho_bruto}")
            calc = processar_bases_excel(caminho_bruto)

            nome_saida = f_bruto[0].replace('.xlsx', '_PRONTO.xlsx')
            caminho_saida = os.path.join(path_ent_at, nome_saida)
            molde_src = os.path.join(path_ent_ant, f_molde[0])
            print(f"  [INFO] Copiando molde: {molde_src} -> {caminho_saida}")

            shutil.copy2(molde_src, caminho_saida)
            copiar_abas_brutas(
                destino_xlsx=caminho_saida,
                fonte_xlsx=caminho_bruto,
                padroes=(r'^MENSAL', r'^COPART', r'^MNC.*20'),
                preferir_xlwings=True,
            )
            preencher_layout(caminho_saida, calc, nfs, MES_ANTERIOR_NOME, MES_ATUAL_NOME, VENCIMENTO_PADRAO)
            print(f"  [OK] Planilha gerada: {nome_saida}")

            if GERAR_AUDITORIA:
                rel = auditar_planilha(caminho_saida, nfs, entidade=ent)
                if rel:
                    relatorios.append(rel)

        except Exception as e:
            print(f"  [ERRO] {ent}: {e}")
            traceback.print_exc()

    return relatorios


# ===== Exportação do resumo de auditoria =====
def gerar_resumo_auditoria_xlsx(relatorios: list, destino_path: str):
    if not relatorios:
        print("[RESUMO] Nada para exportar.")
        return None

    df_nf = pd.concat([r['por_nf'] for r in relatorios], ignore_index=True) if relatorios else pd.DataFrame()
    df_ct = pd.concat([r['por_contrato'] for r in relatorios], ignore_index=True) if relatorios else pd.DataFrame()

    if not df_nf.empty:
        df_nf.sort_values(['ENTIDADE', 'CONTRATO', 'Nº NOTA (pdf)'], inplace=True, ignore_index=True)
    if not df_ct.empty:
        df_ct.sort_values(['ENTIDADE', 'CONTRATO'], inplace=True, ignore_index=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'AUDITORIA_POR_NF'
    ws2 = wb.create_sheet('AUDITORIA_POR_CONTRATO')

    header_fill = PatternFill(start_color='FFEEEEEE', end_color='FFEEEEEE', fill_type='solid')
    bold = Font(bold=True)
    thin = Side(style='thin', color='FF888888')
    border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
    cur_fmt = 'R$ #,##0.00'

    def _write_df(ws, df: pd.DataFrame, money_cols: list[str]):
        ws.append(list(df.columns))
        for j in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=j)
            c.font = bold
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center')
            c.border = border_all

        for _, row in df.iterrows():
            ws.append([row[c] for c in df.columns])

        for i in range(2, ws.max_row + 1):
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=i, column=col_idx)
                cell.border = border_all
                if col_name in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = cur_fmt
                if isinstance(cell.value, str) and len(cell.value) > 40:
                    cell.alignment = Alignment(wrap_text=True)

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            for r in range(2, ws.max_row + 1):
                try:
                    max_len = max(max_len, len(str(ws.cell(row=r, column=col_idx).value or '')))
                except Exception:
                    pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))
        ws.freeze_panes = 'A2'

    money_nf = ['VALOR_SERVICO(pdf)', 'BRUTO(layout)', 'DIF_BRUTO(pdf_vs_layout)']
    money_ct = ['BRUTO(pdf)_contrato', 'BRUTO(layout)_contrato', 'DIF_CONTRATO(pdf_vs_layout)']

    if not df_nf.empty:
        _write_df(ws1, df_nf, money_nf)
    if not df_ct.empty:
        _write_df(ws2, df_ct, money_ct)

    pasta = os.path.dirname(destino_path)
    if pasta and not os.path.isdir(pasta):
        os.makedirs(pasta, exist_ok=True)
    wb.save(destino_path)
    print(f"[OK] Resumo de auditoria salvo em: {destino_path}")
    return destino_path


# ===== Cópia das abas brutas =====
def copiar_abas_brutas(destino_xlsx: str, fonte_xlsx: str,
                        padroes=(r'^MENSAL', r'^COPART', r'^MNC.*20'),
                        preferir_xlwings=True):
    """Copia abas do bruto para o layout de destino, preservando formatação."""
    wb_src_probe = openpyxl.load_workbook(fonte_xlsx, read_only=True, data_only=True)
    nomes_src = wb_src_probe.sheetnames
    abas_match = []
    for rx in padroes:
        rxc = re.compile(rx, re.IGNORECASE)
        m = next((n for n in nomes_src if rxc.search(n)), None)
        if m:
            abas_match.append(m)
        else:
            print(f"[AVISO] Aba não encontrada para padrão: {rx}")
    wb_src_probe.close()

    if preferir_xlwings:
        try:
            import xlwings as xw
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False
            try:
                wb_dst = xw.Book(destino_xlsx)
                wb_src = xw.Book(fonte_xlsx)
                for nome in abas_match:
                    for s in list(wb_dst.sheets):
                        if s.name.strip().lower() == nome.strip().lower():
                            s.delete()
                    src_sheet = wb_src.sheets[nome]
                    src_sheet.api.Copy(After=wb_dst.sheets[-1].api)
                    new_sheet = wb_dst.sheets[-1]
                    try:
                        new_sheet.name = nome
                    except Exception:
                        i = 1
                        while True:
                            candidate = f"{nome}_copiado_{i}"
                            try:
                                new_sheet.name = candidate
                                break
                            except Exception:
                                i += 1
                wb_src.close()
                wb_dst.save()
                wb_dst.close()
                print("[OK] Abas copiadas com xlwings.")
                return
            finally:
                app.quit()
        except Exception as e:
            print(f"[FALLBACK] xlwings indisponível ({e}). Usando openpyxl...")

    from openpyxl.styles import (
        Font as _Font, PatternFill as _PatternFill,
        Alignment as _Alignment, Border as _Border, Side as _Side,
    )

    def _safe_side(s):
        try:
            return _Side(style=getattr(s, "style", None), color=getattr(s, "color", None))
        except Exception:
            return _Side()

    def _safe_border(b):
        try:
            return _Border(
                left=_safe_side(b.left), right=_safe_side(b.right),
                top=_safe_side(b.top), bottom=_safe_side(b.bottom),
                diagonal=_safe_side(b.diagonal), vertical=_safe_side(b.vertical),
                horizontal=_safe_side(b.horizontal),
            )
        except Exception:
            return _Border()

    def _safe_font(f):
        try:
            return _Font(
                name=getattr(f, "name", None), size=getattr(f, "size", None),
                bold=getattr(f, "bold", None), italic=getattr(f, "italic", None),
                vertAlign=getattr(f, "vertAlign", None), underline=getattr(f, "underline", None),
                strike=getattr(f, "strike", None), color=getattr(f, "color", None),
            )
        except Exception:
            return _Font()

    def _safe_fill(fl):
        try:
            return _PatternFill(
                fill_type=getattr(fl, "fill_type", None),
                fgColor=getattr(fl, "fgColor", None),
                bgColor=getattr(fl, "bgColor", None),
            )
        except Exception:
            return _PatternFill()

    def _safe_alignment(a):
        try:
            return _Alignment(
                horizontal=getattr(a, "horizontal", None),
                vertical=getattr(a, "vertical", None),
                text_rotation=getattr(a, "text_rotation", 0),
                wrap_text=getattr(a, "wrap_text", None),
                shrink_to_fit=getattr(a, "shrink_to_fit", None),
                indent=getattr(a, "indent", 0),
            )
        except Exception:
            return _Alignment()

    def _copy_sheet_with_styles(ws_o, ws_d):
        for row in ws_o.iter_rows():
            ws_d.append([c.value for c in row])
        for r in ws_o.iter_rows():
            for c in r:
                d = ws_d.cell(row=c.row, column=c.column)
                try:
                    if c.has_style:
                        d.font = _safe_font(c.font)
                        d.fill = _safe_fill(c.fill)
                        d.alignment = _safe_alignment(c.alignment)
                        d.border = _safe_border(c.border)
                        d.number_format = c.number_format
                except Exception:
                    pass
        for k, cd in ws_o.column_dimensions.items():
            try:
                ws_d.column_dimensions[k].width = cd.width
            except Exception:
                pass
        for k, rd in ws_o.row_dimensions.items():
            try:
                ws_d.row_dimensions[int(k)].height = rd.height
            except Exception:
                pass
        try:
            for rng in ws_o.merged_cells.ranges:
                ws_d.merge_cells(str(rng))
        except Exception:
            pass
        ws_d.freeze_panes = ws_o.freeze_panes
        ws_d.sheet_format = ws_o.sheet_format
        ws_d.page_margins = ws_o.page_margins
        ws_d.page_setup = ws_o.page_setup
        ws_d.print_options = ws_o.print_options

    src = openpyxl.load_workbook(fonte_xlsx, data_only=False, read_only=False)
    dst = openpyxl.load_workbook(destino_xlsx)

    for nome in abas_match:
        if nome in dst.sheetnames:
            idx = dst.sheetnames.index(nome)
            dst.remove(dst[nome])
            ws_d = dst.create_sheet(nome, idx)
        else:
            ws_d = dst.create_sheet(nome)
        _copy_sheet_with_styles(src[nome], ws_d)

    try:
        dst.save(destino_xlsx)
    except PermissionError:
        base, ext = os.path.splitext(destino_xlsx)
        alt = f"{base}_UNLOCK{ext}"
        dst.save(alt)
        print(f"[AVISO] Arquivo bloqueado; salvo como '{os.path.basename(alt)}'.")

    src.close()
    print("[OK] Abas copiadas com openpyxl.")


if __name__ == "__main__":
    rels = executar()
    try:
        print("\n[FINALIZADO] Processamento concluído.")
        op = input("Digite [E] para exportar resumo XLSX da auditoria, ou ENTER para sair: ").strip().upper()
        if op == 'E':
            hoje = datetime.datetime.now().strftime('%Y-%m-%d_%Hh%M')
            destino = os.path.join(PASTA_MES_ATUAL, f"RESUMO_AUDITORIA_{hoje}.xlsx")
            gerar_resumo_auditoria_xlsx(rels, destino)
    finally:
        input("\nPressione ENTER para fechar… ")
